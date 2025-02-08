# %%
import pandas as pd

# 读取 Excel 文件并将数据存储到 DataFrame 中
drgs_raw = pd.read_excel('2024data.xlsx')

# 打印 DataFrame 的前几行以验证数据是否正确导入
print(drgs_raw.head())


# %%

# 假设 DRG 分组列名为 'DRG分组'
# 统计不同 DRG 分组的结余总和

# 统计不同 DRG 分组的结余均值、例数和结算金额的平均数
commonDRGsGroups = drgs_raw.groupby('DRG分组').agg(
    结余均值=('结余', 'mean'),
    例数=('结余', 'count'),
    结算金额=('结算金额', 'mean')
)

# 计算每个 DRG 分组中指定列的中位数总和
base_cost_columns = ['其他费', '检查费', '中药费', '麻醉材料费', '化验费', '伙食费', '麻醉药品费', '护理费', '诊疗费', '治疗费', '床位费', '麻醉其他费']
base_costs = drgs_raw.groupby('DRG分组')[base_cost_columns].mean().sum(axis=1)

# 将中位数总和添加到 commonDRGsGroups 中作为新的一列
commonDRGsGroups['基础费用均值'] = base_costs

# 增加一列 剩余额度 = 结算金额 - 基础费用
commonDRGsGroups['剩余额度'] = commonDRGsGroups['结算金额'] - commonDRGsGroups['基础费用均值']

# 根据例数逆序排序
commonDRGsGroups = commonDRGsGroups.sort_values(by='例数', ascending=False)

# 将结余均值、例数、结算金额平均数和不可控费用转换为整数
commonDRGsGroups['结余均值'] = commonDRGsGroups['结余均值'].astype(int)
commonDRGsGroups['例数'] = commonDRGsGroups['例数'].astype(int)
commonDRGsGroups['结算金额'] = commonDRGsGroups['结算金额'].astype(int)
commonDRGsGroups['基础费用均值'] = commonDRGsGroups['基础费用均值'].astype(int)
commonDRGsGroups['剩余额度'] = commonDRGsGroups['剩余额度'].astype(int)

# 删除例数小于 20 的行
commonDRGsGroups = commonDRGsGroups[commonDRGsGroups['例数'] >= 20]

# 保存commonDRGsGroups到Excel文件
commonDRGsGroups.to_excel('commonDRGsGroups.xlsx', index=True)

# %%

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

# 遍历 commonDRGsGroups 的索引（即 DRG 分组）
excel_path = 'grouped_stats_summary.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    for drg_group in commonDRGsGroups.index:
        # 筛选出当前 DRG 分组的数据
        filtered_df = drgs_raw[drgs_raw['DRG分组'] == drg_group]

        # 根据 主要手术 和 主要诊断 分组，统计结余均值、例数、材料费均值、手术费均值、西药费均值、中成药费均值和住院号列表
        grouped_stats = filtered_df.groupby(['主要手术', '主要诊断']).agg(
            结余均值=('结余', 'mean'),
            例数=('结余', 'count'),
            手术费均值=('手术费', 'mean'),
            材料费均值=('材料费', 'mean'),
            西药费均值=('西药费', 'mean'),
            中成药费均值=('中成药费', 'mean')  # 新增中成药费均值统计
        )

        # 将结余均值、例数、材料费均值、手术费均值、西药费均值和中成药费均值转换为整数
        grouped_stats['结余均值'] = grouped_stats['结余均值'].astype(int)
        grouped_stats['例数'] = grouped_stats['例数'].astype(int)
        grouped_stats['手术费均值'] = grouped_stats['手术费均值'].astype(int)
        grouped_stats['材料费均值'] = grouped_stats['材料费均值'].astype(int)
        grouped_stats['西药费均值'] = grouped_stats['西药费均值'].astype(int)
        grouped_stats['中成药费均值'] = grouped_stats['中成药费均值'].astype(int)  # 新增中成药费均值转换为整数

        # 删除例数小于5的行，如果所有的例数都小于5，则保留所有的行
        if (grouped_stats['例数'] < 5).all():
            pass
        else:
            grouped_stats = grouped_stats[grouped_stats['例数'] >= 5]

        # 按例数降序排名
        grouped_stats = grouped_stats.sort_values(by='例数', ascending=False)

        # 检查 grouped_stats 是否为空
        if grouped_stats.empty:
            continue

        # 保存到 Excel 文件的不同分页
        sheet_name = drg_group.replace('/', '_')[:31]  # 限制表名长度不超过 31 个字符
        grouped_stats.to_excel(writer, sheet_name=sheet_name)

        # 获取当前工作表
        worksheet = writer.sheets[sheet_name]

        # 增加第一列和第二列的宽度
        for col in [1, 2]:
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 45  # 可以根据需要调整宽度值

        for col in [3, 4, 5, 6, 7, 8]:  # 新增一列，调整列数
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 10  # 可以根据需要调整宽度值

        # 设置页面方向为横向
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

        # 设置斑马纹背景
        fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        fill_header = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # 设置表头背景色和字体
        small_font = Font(size=10)
        for col in range(1, len(grouped_stats.columns) + 3):  # +2 是因为有索引列
            cell = worksheet.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = small_font

        # 设置偶数行背景色和字体
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = small_font
                if row % 2 == 0:
                    cell.fill = fill_even

        # 获取当前 DRG 分组的结算金额、结余、例数、基础费用和剩余额度
        # Define alignment styles
        align_right = Alignment(horizontal='right')
        align_left = Alignment(horizontal='left')
        
        # Get the settlement amount, balance, number of cases, base cost and remaining quota for the current DRG group
        settlement_amount = commonDRGsGroups.loc[drg_group, '结算金额']
        balance = commonDRGsGroups.loc[drg_group, '结余均值']
        cases = commonDRGsGroups.loc[drg_group, '例数']
        base_cost = commonDRGsGroups.loc[drg_group, '基础费用均值']
        remaining_quota = commonDRGsGroups.loc[drg_group, '剩余额度']
        
        # Add summary information at the end of the table
        last_row = worksheet.max_row + 2  # Add summary information after leaving a blank row
        # Set the first row to bold
        bold_small_font = Font(size=10, bold=True)
        worksheet.cell(row=last_row, column=1, value=f'当前 DRG 分组: {drg_group}').font = bold_small_font
        
        # Write the settlement amount
        worksheet.cell(row=last_row + 1, column=1, value='DRGS结算金额:').font = small_font
        worksheet.cell(row=last_row + 1, column=1).alignment = align_right
        worksheet.cell(row=last_row + 1, column=2, value=settlement_amount).font = small_font
        worksheet.cell(row=last_row + 1, column=2).alignment = align_left
        
        # Write the balance
        worksheet.cell(row=last_row + 2, column=1, value='本组平均结余:').font = small_font
        worksheet.cell(row=last_row + 2, column=1).alignment = align_right
        worksheet.cell(row=last_row + 2, column=2, value=balance).font = small_font
        worksheet.cell(row=last_row + 2, column=2).alignment = align_left
        
        # Write the number of cases
        worksheet.cell(row=last_row + 3, column=1, value='本组病例数:').font = small_font
        worksheet.cell(row=last_row + 3, column=1).alignment = align_right
        worksheet.cell(row=last_row + 3, column=2, value=cases).font = small_font
        worksheet.cell(row=last_row + 3, column=2).alignment = align_left

        # Write the base cost
        worksheet.cell(row=last_row + 4, column=1, value='基础费用均值:').font = small_font
        worksheet.cell(row=last_row + 4, column=1).alignment = align_right
        worksheet.cell(row=last_row + 4, column=2, value=base_cost).font = small_font
        worksheet.cell(row=last_row + 4, column=2).alignment = align_left

        # Write the remaining quota
        worksheet.cell(row=last_row + 5, column=1, value='本组剩余额度（手术费、材料费、药费）:').font = small_font
        worksheet.cell(row=last_row + 5, column=1).alignment = align_right
        worksheet.cell(row=last_row + 5, column=2, value=remaining_quota).font = small_font
        worksheet.cell(row=last_row + 5, column=2).alignment = align_left

        for i in range(1, 4):
            worksheet.cell(row=last_row + i, column=1).font = small_font
            worksheet.cell(row=last_row + i, column=1).alignment = align_right
            worksheet.cell(row=last_row + i, column=2).font = small_font
            worksheet.cell(row=last_row + i, column=2).alignment = align_left

        # 设置三线表格样式
        thin = Side(border_style='thin', color='FF000000')
        header_border = Border(bottom=thin)
        table_top_border = Border(top=thin)
        table_bottom_border = Border(bottom=thin)

        # 设置表头下边框
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=1, column=col).border = header_border

        # 设置表格上边框
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=2, column=col).border = table_top_border

        # 设置表格下边框，仅针对 grouped_stats 表格
        table_last_row = last_row - 2
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=table_last_row, column=col).border = table_bottom_border

        # 删除表格中间的边框线
        for row in range(3, table_last_row):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).border = Border()

        # 设置打印缩放为一页
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1

# %%

from collections import Counter

# 筛选出 DRG分组 包含 “伴严重并发症” 的行
filtered_df = drgs_raw[drgs_raw['DRG分组'].str.contains('伴严重并发症', na=False)]

# 提取其他诊断列中的所有诊断，并将其拆分为单个诊断
all_diagnoses = []
for diag_str in filtered_df['其他诊断']:
    if isinstance(diag_str, str):
        diag_list = diag_str.split(',')
        all_diagnoses.extend(diag_list)

# 统计每个诊断的出现频次
diagnosis_counter = Counter(all_diagnoses)

# 按频次降序排名
severComplications = sorted(diagnosis_counter.items(), key=lambda item: item[1], reverse=True)


# %%
import pandas as pd
from collections import Counter

# 筛选出 DRG分组 包含 “不伴并发症” 的行
filtered_df = drgs_raw[drgs_raw['DRG分组'].str.contains('不伴并发症', na=False)]

# 提取其他诊断列中的所有诊断，并将其拆分为单个诊断
all_diagnoses = []
for diag_str in filtered_df['其他诊断']:
    if isinstance(diag_str, str):
        diag_list = diag_str.split(',')
        all_diagnoses.extend(diag_list)

# 统计每个诊断的出现频次
diagnosis_counter = Counter(all_diagnoses)

# 按频次降序排名
sorted_diagnoses = sorted(diagnosis_counter.items(), key=lambda item: item[1], reverse=True)

# 创建 DataFrame
withoutComplications = pd.DataFrame(sorted_diagnoses, columns=['诊断', '频次'])



from collections import Counter

# 筛选出 DRG分组 包含 “伴一般并发症” 的行
filtered_df = drgs_raw[drgs_raw['DRG分组'].str.contains('伴一般并发症', na=False)]

# 提取其他诊断列中的所有诊断，并将其拆分为单个诊断
all_diagnoses = []
for diag_str in filtered_df['其他诊断']:
    if isinstance(diag_str, str):
        diag_list = diag_str.split(',')
        all_diagnoses.extend(diag_list)

# 统计每个诊断的出现频次
diagnosis_counter = Counter(all_diagnoses)

# 按频次降序排名
generalComplications = sorted(diagnosis_counter.items(), key=lambda item: item[1], reverse=True)

generalComplications = pd.DataFrame(generalComplications, columns=['诊断', '频次'])

from collections import Counter

# 筛选出 DRG分组 包含 “伴严重并发症” 的行
filtered_df = drgs_raw[drgs_raw['DRG分组'].str.contains('伴严重并发症', na=False)]

# 提取其他诊断列中的所有诊断，并将其拆分为单个诊断
all_diagnoses = []
for diag_str in filtered_df['其他诊断']:
    if isinstance(diag_str, str):
        diag_list = diag_str.split(',')
        all_diagnoses.extend(diag_list)

# 统计每个诊断的出现频次
diagnosis_counter = Counter(all_diagnoses)

# 按频次降序排名
severComplications = sorted(diagnosis_counter.items(), key=lambda item: item[1], reverse=True)

severComplications = pd.DataFrame(severComplications, columns=['诊断', '频次'])

# %%

# 筛选出 generalComplications 中的诊断除去 withoutComplications 中的诊断
filtered_generalComplications = generalComplications[~generalComplications['诊断'].isin(withoutComplications['诊断'])]
print(filtered_generalComplications)

# %%
# 筛选出 severComplications 中的诊断除去 withoutComplications 和 generalComplications 中的诊断
filtered_severComplications = severComplications[~severComplications['诊断'].isin(withoutComplications['诊断']) & ~severComplications['诊断'].isin(generalComplications['诊断'])]
print(filtered_severComplications)
# %%
